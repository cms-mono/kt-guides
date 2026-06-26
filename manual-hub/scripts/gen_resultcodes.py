# -*- coding: utf-8 -*-
import openpyxl, json, io, glob, re

SRC = r"archive/창용's/통합_에러코드.xlsx"
# RCS 보강 원본: 공식 KT RCS ErrorCode 워크북(여러 버전 중 최신 자동 선택)
RCS_SRC_GLOB = r"archive/창용's/RCS/*ErrorCode*.xlsx"
OUT = r"26년 KPI/4. 문서 업데이트/웹매뉴얼/손팀장님/통합 메뉴얼 목업/manual-hub/resultcodes.js"

# 엑셀 서비스명 -> 그룹 정의 (사용자 요청 5개)
GROUPS = [
    ("mcs-agent",  "KT 스마트메시지 (Agent)",   "스마트메시지 Agent", {"KT스마트메시지 (Agent)"}),
    ("mcs-openapi","KT 스마트메시지 (OpenAPI)", "스마트메시지 OpenAPI", {"KT스마트메시지 (openAPI)"}),
    ("communis",   "커뮤니즈 (Communis)",       "커뮤니즈", {"Communis API"}),
    ("rcs",        "RCS",                       "RCS", {"RCS"}),
    ("alimtalk",   "알림톡",                    "알림톡", {"알림톡"}),
]

def norm(v):
    return ("" if v is None else str(v)).replace("\r\n", "\n").strip()

# 외부 고객에게 노출되면 안 되는 KT 내부 안내(운용실/운영실 직접 연락)를
# '모노커뮤니케이션즈를 통한 문의'로 치환한다. (긴 표현 먼저 — 부분치환 방지)
INTERNAL_REPLACEMENTS = [
    ("KT 운용실로부터", "모노커뮤니케이션즈를 통해"),
    ("KT운용실로부터", "모노커뮤니케이션즈를 통해"),
    ("KT 운용실에", "모노커뮤니케이션즈에"),
    ("KT운용실에", "모노커뮤니케이션즈에"),
    ("KT 운용실", "모노커뮤니케이션즈"),
    ("KT운용실", "모노커뮤니케이션즈"),
    ("운영실로 차단 해제 요청이 가능함", "모노커뮤니케이션즈를 통해 차단 해제 요청이 가능함"),
    ("운영실로", "모노커뮤니케이션즈로"),
    ("운영실에", "모노커뮤니케이션즈에"),
    ("운영실", "모노커뮤니케이션즈"),
]
def sanitize(s):
    if not s:
        return s
    for a, b in INTERNAL_REPLACEMENTS:
        s = s.replace(a, b)
    return s

def svc_norm(v):
    # 공백/줄바꿈 차이 흡수
    return norm(v).replace("\n", " ").replace("  ", " ").strip()

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(min_row=2, values_only=True))

buckets = {gid: [] for gid, _, _, _ in GROUPS}
svc_to_gid = {}
for gid, _, _, names in GROUPS:
    for n in names:
        svc_to_gid[n] = gid

leftover = {}
for r in rows:
    code = norm(r[1]); title = sanitize(norm(r[2])); desc = sanitize(norm(r[3])); svc = svc_norm(r[4])
    if not svc:
        continue
    if code == "" and title == "" and desc == "":
        continue
    gid = svc_to_gid.get(svc)
    if gid is None:
        leftover[svc] = leftover.get(svc, 0) + 1
        continue
    buckets[gid].append({"code": code, "title": title, "desc": desc})

groups_out = []
for gid, name, short, _ in GROUPS:
    items = buckets[gid]
    groups_out.append({"id": gid, "name": name, "short": short, "count": len(items), "codes": items})

# ── RCS 보강: 공식 RCS ErrorCode 워크북에서 누락 코드 병합 ──────────────
#   통합_에러코드.xlsx의 RCS는 MaaP-FE(50000s)+KT중계(70000s) 부분집합이라 누락 코드가
#   생긴다. 최신 RCS ErrorCode 워크북의 전 범위를 읽어 빠진 코드를 채운다.
#   제목은 기존 RCS와 동일하게 워크북의 '설명' 열(없으면 영문 Message)을 사용한다.
#   각 항목: (시트, 코드열, [제목열 우선순위], lo, hi) — 0-based 열 인덱스
#   ※ 같은 코드가 여러 시트에 있으면 먼저 나온 시트가 우선(아래 순서대로 채움).
RCS_MERGE_RANGES = [
    ("삼성MaaP-GW2.0",  3, [5, 4], 40000, 49999),  # 삼성 MaaP-Core(현행) — 설명(F)→Text(E)
    ("(구)삼성MaaP-GW", 2, [4],    40000, 49999),  # 레거시 — 2.0에 없는 코드만 보충
    ("MaaP-FE",         2, [5],    50000, 59999),  # 이통사 MaaP-FE — 설명
    ("RcsBizCenter",    3, [5],    60000, 69999),  # 이통사 RcsBizCenter — 코드=TO BE, 설명
    ("KT 중계사",       2, [7],    70000, 79999),  # KT 중계 — 설명
]
def _is_code(v):
    return v is not None and bool(re.fullmatch(r"\d{4,5}", str(v).strip()))
def _ver_key(p):
    m = re.search(r"v(\d+)\.(\d+)", p)
    return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
def _title(row, cols):
    for ci in cols:
        if len(row) > ci:
            t = norm(row[ci])
            if t and t != "-":
                return t
    return ""

rcs_added = []
rcs_src_used = None
rcs_files = glob.glob(RCS_SRC_GLOB)
if rcs_files:
    rcs_src_used = sorted(rcs_files, key=_ver_key)[-1]
    wb2 = openpyxl.load_workbook(rcs_src_used, data_only=True)
    rcs_group = next(g for g in groups_out if g["id"] == "rcs")
    existing = {str(c["code"]).strip() for c in rcs_group["codes"]}
    existing.add("00000")  # 기존 '0'(성공)과 중복이므로 추가하지 않음
    for sheet, ci, tcols, lo, hi in RCS_MERGE_RANGES:
        ws2 = wb2[sheet]
        for row in ws2.iter_rows(values_only=True):
            if len(row) <= ci or not _is_code(row[ci]):
                continue
            code = str(row[ci]).strip()
            if not (lo <= int(code) <= hi) or code in existing:
                continue
            existing.add(code)
            rcs_group["codes"].append({"code": code, "title": sanitize(_title(row, tcols)), "desc": ""})
            rcs_added.append(code)
    # 병합 후 RCS 그룹을 코드 오름차순으로 정리(누락분이 제 위치에 놓이도록)
    rcs_group["codes"].sort(key=lambda c: int(c["code"]) if str(c["code"]).strip().isdigit() else 10**9)
    rcs_group["count"] = len(rcs_group["codes"])

data = {
    "updated": "2026-06-18",
    "source": "통합_에러코드.xlsx + KT RCS_ErrorCode v2.83 (㈜다이얼로그스페이스 외)",
    "note": "각 서비스 매뉴얼의 결과코드(응답코드)를 한곳에 모았습니다. 동일 코드라도 서비스마다 의미가 다를 수 있으니 서비스 탭을 선택해 조회하세요.",
    "groups": groups_out,
}

body = "window.HUB_RESULTCODES = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
header = (
    "/* ============================================================\n"
    "   서비스별 결과코드(응답코드) — 단일 출처 데이터\n"
    "   ------------------------------------------------------------\n"
    "   원본: archive/창용's/통합_에러코드.xlsx\n"
    "        + archive/창용's/RCS/*ErrorCode*.xlsx (RCS 누락코드 보강, 최신버전 자동선택)\n"
    "   생성: manual-hub/scripts/gen_resultcodes.py (저장소 루트에서 실행)\n"
    "        → 엑셀 갱신 시 재실행. 이 파일은 자동 생성물이므로 수기 편집 금지.\n"
    "   그룹: KT스마트메시지(Agent) / KT스마트메시지(OpenAPI) /\n"
    "         커뮤니즈(Communis) / RCS / 알림톡\n"
    "   ============================================================ */\n"
)
with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(header + body)

print("WROTE", OUT)
for g in groups_out:
    print(f"  {g['id']:12s} {g['count']:5d}  {g['name']}")
print("LEFTOVER (그룹 미지정):", leftover)
if rcs_src_used:
    print(f"RCS 보강 원본: {rcs_src_used}")
    print(f"RCS 추가 코드 {len(rcs_added)}개: {', '.join(rcs_added)}")
